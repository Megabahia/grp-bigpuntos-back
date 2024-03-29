from apps.CENTRAL.central_catalogo.models import Catalogo
from .models import CreditoPersonas, CodigoCreditoPreaprobado
from apps.PERSONAS.personas_personas.models import Personas
from apps.CORP.corp_empresas.models import Empresas
from apps.PERSONAS.personas_personas.serializers import PersonasSearchSerializer
from .serializers import (
    CreditoPersonasSerializer, CreditoPersonasPersonaSerializer
)
# Publicar en sns
from apps.CORP.corp_creditoPersonas.producer import publish
# Consumir en sqs
from .consumer import get_queue_url

# include json library
import json

from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.conf import settings
# Swagger
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
# excel
import openpyxl
# Generar codigos aleatorios
import string
import random
# Sumar minutos
from dateutil.relativedelta import relativedelta
# ObjectId
from bson import ObjectId
# Lectura de AWS s3
import boto3
import re
from apps.config import config
# logs
from apps.CENTRAL.central_logs.methods import createLog, datosTipoLog, datosProductosMDP
# IMPORTAR ENVIO CONFIGURACION CORREO
from apps.config.util2 import sendEmail

# declaracion variables log
datosAux = datosProductosMDP()
datosTipoLogAux = datosTipoLog()
# asignacion datos modulo
logModulo = datosAux['modulo']
logApi = datosAux['api']
# asignacion tipo de datos
logTransaccion = datosTipoLogAux['transaccion']
logExcepcion = datosTipoLogAux['excepcion']


# CRUD
# CREAR
# 'methods' can be used to apply the same modification to multiple methods
@swagger_auto_schema(methods=['post'], request_body=CreditoPersonasSerializer)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_create(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            request.data['created_at'] = str(timezone_now)
            if 'updated_at' in request.data:
                request.data.pop('updated_at')

            if 'nombres' in request.data:
                if request.data['nombres'] != "":
                    request.data['nombresCompleto'] = request.data['nombres'] + ' ' + request.data['apellidos']

            serializer = CreditoPersonasSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                if 'tipoCredito' in serializer.data and serializer.data['tipoCredito'] == 'null':
                    usuario = serializer.data['user']
                    nombreGarante = usuario['garante']['nombres'] + ' ' + usuario['garante']['apellidos']
                    nombreSolicitante = usuario['nombres'] + ' ' + usuario['apellidos']
                    enviarCorreoSolicitudGarante(usuario['garante']['correoGarante'], serializer.data['_id'], nombreGarante, nombreSolicitante)
                createLog(logModel, serializer.data, logTransaccion)
                # Crear objeto en firebase para las notificaciones
                config.FIREBASE_DB.collection('creditosPersonas').document(serializer.data['_id']).set(serializer.data)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ENCONTRAR UNO
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def creditoPersonas_listOne(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            serializer = CreditoPersonasSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ACTUALIZAR
# 'methods' can be used to apply the same modification to multiple methods
@swagger_auto_schema(methods=['post'], request_body=CreditoPersonasSerializer)
@api_view(['POST'])
# @permission_classes([IsAuthenticated])
def creditoPersonas_update(request, pk):
    request.POST._mutable = True
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'update/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'ESCRIBIR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            logModel['dataEnviada'] = str(request.data)
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).first()
        except CreditoPersonas.DoesNotExist:
            errorNoExiste = {'error': 'No existe'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method == 'POST':
            now = timezone.localtime(timezone.now())
            request.data['updated_at'] = str(now)
            if 'created_at' in request.data:
                request.data.pop('created_at')
            if query.estado == 'Por completar':
                request.data['estado'] = 'Completado'
            if 'tipoCredito' in request.data:
                if request.data['tipoCredito'] == '':
                    request.data['tipoCredito'] = query.canal
            if query.enviado == 0:
                request.data['enviado'] = 1
            if query.alcance is None:
                request.data['alcance'] = 'OMNIGLOBAL'
            serializer = CreditoPersonasSerializer(query, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                if "estado" in request.data:
                    if request.data["estado"] == 'Enviado':
                        # Se envia a la cola de bigpuntos
                        publish(serializer.data)
                        usuario = serializer.data['user']
                        if usuario:
                            email = usuario['email']
                        else:
                            email = serializer.data['empresaInfo']['correo']
                        if 'Pymes' in serializer.data['tipoCredito']:
                            enviarCorreoSolicitudEnviadaLineaCredito(email)
                        else:
                            enviarCorreoSolicitudEnviada(email)
                    if request.data["estado"] == 'Completado':
                        # Se envia a la cola de bigpuntos
                        publish(serializer.data)
                return Response(serializer.data)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)

    # ELIMINAR


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def creditoPersonas_delete(request, pk):
    nowDate = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'delete/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'BORRAR',
        'fechaInicio': str(nowDate),
        'dataEnviada': '{}',
        'fechaFin': str(nowDate),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(user_id=pk, state=1).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
            return Response(status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'DELETE':
            serializer = CreditoPersonasSerializer(query, data={'state': '0', 'updated_at': str(nowDate)}, partial=True)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_200_OK)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoPersonas_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}

            if "empresaComercial_id" in request.data:
                if request.data["empresaComercial_id"] != '':
                    filters['empresaComercial_id'] = ObjectId(request.data["empresaComercial_id"])

            if "empresaIfis_id" in request.data:
                if request.data["empresaIfis_id"] != '':
                    filters['empresaIfis_id'] = ObjectId(request.data["empresaIfis_id"])

            if "estado" in request.data:
                if request.data["estado"] != '':
                    filters['estado__in'] = request.data["estado"]

            if "tipoCredito" in request.data:
                if request.data["tipoCredito"] != '':
                    filters['tipoCredito'] = str(request.data["tipoCredito"])

            if "user_id" in request.data:
                if request.data["user_id"] != '':
                    filters['user_id'] = str(request.data["user_id"])

            if "canal" in request.data:
                if request.data["canal"] != '':
                    filters['canal'] = str(request.data["canal"])

            if "cargarOrigen" in request.data:
                if request.data["cargarOrigen"] != '':
                    filters['cargarOrigen'] = str(request.data["cargarOrigen"])

            # Serializar los datos
            query = CreditoPersonas.objects.filter(**filters).order_by('-created_at')
            serializer = CreditoPersonasSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# METODO SUBIR ARCHIVOS EXCEL
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_creditosPreaprobados(request):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Clientes"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 7:
                    resultadoInsertar = insertarDato_creditoPreaprobado(dato, request.data['empresa_financiera'])
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_creditoPreaprobado(dato, empresa_financiera):
    try:
        timezone_now = timezone.localtime(timezone.now())
        data = {}
        data['vigencia'] = dato[0].replace('"', "")[0:10] if dato[0] != "NULL" else None
        data['concepto'] = dato[1].replace('"', "") if dato[1] != "NULL" else None
        data['monto'] = dato[2].replace('"', "") if dato[2] != "NULL" else None
        data['plazo'] = dato[3].replace('"', "") if dato[3] != "NULL" else None
        data['interes'] = dato[4].replace('"', "") if dato[4] != "NULL" else None
        data['estado'] = 'PreAprobado'
        data['tipoCredito'] = 'PreAprobado'
        data['canal'] = 'PreAprobado'
        persona = Personas.objects.filter(identificacion=dato[5], state=1).first()
        data['user_id'] = persona.user_id
        data['numeroIdentificacion'] = dato[5]
        data['nombres'] = persona.nombres
        data['apellidos'] = persona.apellidos
        data['nombresCompleto'] = persona.nombres + ' ' + persona.apellidos
        data['empresaIfis_id'] = empresa_financiera
        data['empresasAplican'] = dato[6]
        data['created_at'] = str(timezone_now)
        # inserto el dato con los campos requeridos
        CreditoPersonas.objects.create(**data)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


# METODO SUBIR ARCHIVOS EXCEL
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_creditosPreaprobados_empleados(request):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Clientes"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 11:
                    resultadoInsertar = insertarDato_creditoPreaprobado_empleado(dato,
                                                                                 request.data['empresa_financiera'])
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_creditoPreaprobado_empleado(dato, empresa_financiera):
    try:
        timezone_now = timezone.localtime(timezone.now())
        data = {}
        data['vigencia'] = dato[0].replace('"', "")[0:10] if dato[0] != "NULL" else None
        data['concepto'] = dato[1].replace('"', "") if dato[1] != "NULL" else None
        data['monto'] = dato[2].replace('"', "") if dato[2] != "NULL" else None
        data['plazo'] = dato[3].replace('"', "") if dato[3] != "NULL" else None
        data['interes'] = dato[4].replace('"', "") if dato[4] != "NULL" else None
        data['estado'] = 'PreAprobado'
        data['tipoCredito'] = 'Empleado'
        data['canal'] = 'Empleado'
        persona = Personas.objects.filter(identificacion=dato[5], state=1).first()
        data['user_id'] = persona.user_id
        data['numeroIdentificacion'] = dato[5]
        data['nombres'] = persona.nombres
        data['apellidos'] = persona.apellidos
        data['nombresCompleto'] = persona.nombres + ' ' + persona.apellidos
        data['empresaIfis_id'] = empresa_financiera
        data['empresasAplican'] = dato[10]
        data['created_at'] = str(timezone_now)
        # inserto el dato con los campos requeridos
        CreditoPersonas.objects.create(**data)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


# ENCONTRAR UNO
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def creditoPersonas_listOne_persona(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            serializer = CreditoPersonasPersonaSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ENCONTRAR CODIGO CREDITO PREAPROBADO
@api_view(['POST'])
def creditoPersonas_creditoPreaprobado_codigo(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'creditoPreaprobado/codigo',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # Filtros
            filters = {"state": "1"}

            if "codigo" in request.data:
                if request.data["codigo"] != '':
                    filters['codigoPreaprobado'] = request.data["codigo"]

            if "cedula" in request.data:
                if request.data["cedula"] != '':
                    filters['numeroIdentificacion'] = request.data["cedula"]

            # Serializar los datos
            try:
                query = CreditoPersonas.objects.get(**filters)
            except CreditoPersonas.DoesNotExist:
                err = {"error": "No existe"}
                createLog(logModel, err, logExcepcion)
                return Response(err, status=status.HTTP_404_NOT_FOUND)

            # response = {'monto': query.monto, 'nombreCompleto': query.nombres + ' ' + query.apellidos,
            #             'tipoPersona': query.tipoPersona, 'estadoCivil': query.estadoCivil}
            # query.state = 0
            # query.save()

            # envio de datos
            serializer = CreditoPersonasSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def creditoPersonas_lecturaArchivos(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'lecturaArchivos/' + pk,
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).first()
        except CreditoPersonas.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            print(query.identificacion.name)
            dato1 = None if query.identificacion.name is None else obtenerDatosArchivos(str(query.identificacion.name))
            dato2 = None if query.ruc.name is None else obtenerDatosArchivos(str(query.ruc.name))
            # serializer = CreditoPersonasPersonaSerializer(query)
            # createLog(logModel, serializer.data, logTransaccion)
            return Response({'cedula': dato1, 'ruc': dato2}, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


def obtenerDatosArchivos(nombreArchivo):
    # Function invokes
    jobId = InvokeTextDetectJob('globalredpymes', nombreArchivo)
    print("Started job with id: {}".format(jobId))
    respuesta = {}
    if (CheckJobComplete(jobId)):
        response = JobResults(jobId)
        for resultPage in response:
            for item in resultPage["Blocks"]:
                if item['BlockType'] == 'LINE':
                    if re.match("\d{10}001", item['Text']):
                        respuesta['ruc'] = item['Text']

                    elif re.match("No. \d{9}-[0-9]", item['Text']):
                        respuesta['identificacion'] = item['Text'][4:]

                    elif re.match("^\d{4}\-(0?[1-9]|1[012])\-(0?[1-9]|[12][0-9]|3[01])$", item['Text']):
                        respuesta['fechaExpiracion'] = item['Text']

                    elif re.match("[aA-Zz]\d{4}[aA-Zz]\d{4}", item['Text']):
                        respuesta['codigoDactilar'] = item['Text']

    print("-------------------Imprimir-----------------")
    return respuesta


import time


## Textract APIs used - "start_document_text_detection", "get_document_text_detection"
def InvokeTextDetectJob(bucket, nombreArchivo):
    response = None
    textarctmodule = boto3.client(
        'textract',
        aws_access_key_id=config.AWS_ACCESS_KEY_ID_TEXTRACT,
        aws_secret_access_key=config.AWS_SECRET_ACCESS_KEY_TEXTRACT,
        region_name='us-east-1'
    )
    response = textarctmodule.start_document_text_detection(
        DocumentLocation={
            'S3Object': {
                'Bucket': bucket,
                # 'Name': nombreArchivo
                'Name': 'CORP/documentosCreditosPersonas/62d97613bceeaa781e803920_1658498310065_comprobante_1.pdf'
            }
        }
    )
    return response["JobId"]


def CheckJobComplete(jobId):
    time.sleep(5)
    client = boto3.client(
        'textract',
        aws_access_key_id=config.AWS_ACCESS_KEY_ID_TEXTRACT,
        aws_secret_access_key=config.AWS_SECRET_ACCESS_KEY_TEXTRACT,
        region_name='us-east-1'
    )
    response = client.get_document_text_detection(JobId=jobId)
    status = response["JobStatus"]
    print("Job status: {}".format(status))
    while (status == "IN_PROGRESS"):
        time.sleep(5)
        response = client.get_document_text_detection(JobId=jobId)
        status = response["JobStatus"]
        print("Job status: {}".format(status))
    return status


def JobResults(jobId):
    pages = []
    client = boto3.client(
        'textract',
        aws_access_key_id=config.AWS_ACCESS_KEY_ID_TEXTRACT,
        aws_secret_access_key=config.AWS_SECRET_ACCESS_KEY_TEXTRACT,
        region_name='us-east-1'
    )
    response = client.get_document_text_detection(JobId=jobId)

    pages.append(response)
    print("Resultset page recieved: {}".format(len(pages)))
    nextToken = None
    if ('NextToken' in response):
        nextToken = response['NextToken']
        while (nextToken):
            response = client.get_document_text_detection(JobId=jobId, NextToken=nextToken)
            pages.append(response)
            print("Resultset page recieved: {}".format(len(pages)))
            nextToken = None
            if ('NextToken' in response):
                nextToken = response['NextToken']
    return pages


@api_view(['GET'])
def pruebaConsumer(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        get_queue_url()
        msg = {"msg": "Se actualizo la cola"}

        return Response(msg, status=status.HTTP_202_ACCEPTED)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


def enviarCorreoSolicitud(email):
    subject, from_email, to = 'Generación de código para crédito aprobado', "credicompra.bigpuntos@corporacionomniglobal.com", \
                              email
    txt_content = f"""
                        Su microcrédito línea de crédito para pago a proveedores ha sido NEGADO.
                            Crédito Pagos es la mejor opción para el crecimiento de su negocio
                        Atentamente,
                        Global RedPyme – Crédito Pagos
    """
    html_content = f"""
                <html>
                    <body>
                        <h1>LO SENTIMOS!! </h1>
                        <p>Su microcrédito línea de crédito para pago a proveedores ha sido <b>NEGADO.</b> 
                        Esperamos poder ayudarle en una próxima ocasión.</p>
                        <br>
                        <br>
                        <p>Crédito Pagos es la mejor opción para el crecimiento de su negocio</p>
                        <br>
                        <br>
                        Atentamente,
                        <br>
                        Global RedPyme – Crédito Pagos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


def enviarCorreoSolicitudEnviada(email):
    subject, from_email, to = 'Estamos revisando sus documentos', "credicompra.bigpuntos@corporacionomniglobal.com", \
                              email
    txt_content = f"""
                        REVISIÓN DE DOCUMENTOS EN PROCESO
                        Sus documentos para acceder a un Crédito de consumo otorgado por una Cooperativa de Ahorro y
                            Crédito regulada para realizar compras en los mejores Locales Comerciales del país, 
                            están siendo revisados.
                        Le mantendremos informado a través de nuestros canales.
                        Atentamente,
                        CrediCompra – Big Puntos
    """
    html_content = f"""
                <html>
                    <body>
                        <h1>REVISIÓN DE DOCUMENTOS EN PROCESO</h1>
                        <p>
                            Sus documentos para acceder a un Crédito de consumo otorgado por una Cooperativa de Ahorro y
                            Crédito regulada para realizar compras en los mejores Locales Comerciales del país, 
                            están siendo revisados.
                         </p>
                        <br>
                        <br>
                        <p>Le mantendremos informado a través de nuestros canales.</p>
                        <br>
                        <br>
                        Atentamente,
                        <br>
                        CrediCompra – Big Puntos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


def enviarCorreoSolicitudEnviadaLineaCredito(email):
    subject, from_email, to = 'Estamos revisando sus documentos', "credicompra.bigpuntos@corporacionomniglobal.com", \
                              email
    txt_content = f"""
        REVISIÓN DE DOCUMENTOS EN PROCESO
        Sus documentos para acceder a una Línea de Crédito para realizar pagos a sus proveedores y/o
         empleados, otorgada por una Cooperativa de Ahorro y Crédito regulada, están siendo revisados.

        Le mantendremos informado a través de nuestros canales.

        Atentamente,
        Equipo Global Redpyme – Crédito Pagos
    """
    html_content = f"""
        <html>
            <body>
                <h1>REVISIÓN DE DOCUMENTOS EN PROCESO</h1>
                <p>
                    Sus documentos para acceder a una Línea de Crédito para realizar pagos a sus proveedores 
                    y/o empleados, otorgada por una Cooperativa de Ahorro y Crédito regulada, están siendo revisados.
                 </p>
                <br>
                <br>
                <p>
                Le mantendremos informado a través de nuestros canales.
                </p>
                <br>
                <br>
                Atentamente,
                <br>
                Equipo Global Redpyme – Crédito Pagos
                <br>
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)


# ENCONTRAR UNO
@api_view(['POST'])
def creditoPersonas_listOne_sinAutenticar(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        if request.method == 'POST':
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), tipoCredito=request.data['tipoCredito'],
                                                   state=1).first()
            if query is None:
                err = {"error": "No existe"}
                createLog(logModel, err, logExcepcion)
                return Response(err, status=status.HTTP_404_NOT_FOUND)
            # tomar el dato
            serializer = CreditoPersonasSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ACTUALIZAR SIN AUTENTICAR
@api_view(['POST'])
def creditoPersonas_update_sinAutenticar(request, pk):
    request.POST._mutable = True
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'update/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'ESCRIBIR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            logModel['dataEnviada'] = str(request.data)
            query = CreditoPersonas.objects.filter(pk=ObjectId(pk), state=1).first()
        except CreditoPersonas.DoesNotExist:
            errorNoExiste = {'error': 'No existe'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method == 'POST':
            now = timezone.localtime(timezone.now())
            request.data['updated_at'] = str(now)
            if 'created_at' in request.data:
                request.data.pop('created_at')
            if query.enviado == 0:
                request.data['enviado'] = 1
            if query.alcance is None:
                request.data['alcance'] = 'OMNIGLOBAL'
            serializer = CreditoPersonasSerializer(query, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


def enviarCorreoSolicitudGarante(email, id, garante, solicitante):
    subject, from_email, to = 'Autorización de la Madrina/Padrino', "credicompra.bigpuntos@corporacionomniglobal.com", \
                              email
    txt_content = f"""
        Autorización de Crédito de Consumo
        
        Estimad@ {garante}, {solicitante} desea que usted le apadrine para poder acceder a un Crédito de Consumo para realizar compras.

        Para confirmar su aprobación como Garantía para acceder al Crédito, haga click en el siguiente enlace y confirme sus datos: 
        {config.API_FRONT_END_CENTRAL}/pages/confirmacion-garante/{id}
        
        Atentamente,
        CrediCompra – Big Puntos
    """
    html_content = f"""
        <html>
            <body>
                <h1>
                Autorización de Crédito de Consumo
                </h1>
                <p>
                Estimad@ {garante}, {solicitante} desea que usted le apadrine para poder acceder a un Crédito
                 de Consumo para realizar compras.
                </p>
                <br>
                <p>
                Para confirmar su aprobación como Garantía para acceder al Crédito, haga click en el siguiente enlace
                 y confirme sus datos: <a href='{config.API_FRONT_END_CENTRAL}/pages/confirmacion-garante/{id}'>ENLACE</a>
                </p>
                <br>
                Atentamente,
                <br>
                CrediCompra – Big Puntos
                <br>
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)
