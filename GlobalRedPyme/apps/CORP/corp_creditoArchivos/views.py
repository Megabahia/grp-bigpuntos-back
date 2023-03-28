from .models import PreAprobados
from ...CENTRAL.central_catalogo.models import Catalogo
from ...CORP.corp_creditoPersonas.models import CreditoPersonas
from ...CORP.corp_creditoPersonas.serializers import CreditoPersonasSerializer
from ...PERSONAS.personas_personas.models import Personas
from .serializers import (
    CreditoArchivosSerializer
)
from ...CORP.corp_empresas.models import Empleados
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.conf import settings
# Utils
from ...utils import utils
# Constantes
from .constants import empresaInfo
# Importar boto3
import boto3
import tempfile
import environ
import os
# Generar codigos aleatorios
import string
import random
# Sumar Fechas
from datetime import datetime
from datetime import timedelta
# Enviar Correo
from ...config.util import sendEmail
# excel
import openpyxl
# ObjectId
from bson import ObjectId
# logs
from ...CENTRAL.central_logs.methods import createLog, datosTipoLog, datosProductosMDP
# Importar en producer de los creditos personas
from ..corp_creditoPersonas.producer import publish
# Importar configuraciones
from ...config import config

# declaracion variables log
datosAux = datosProductosMDP()
datosTipoLogAux = datosTipoLog()
# asignacion datos modulo
logModulo = datosAux['modulo']
logApi = datosAux['api']
# asignacion tipo de datos
logTransaccion = datosTipoLogAux['transaccion']
logExcepcion = datosTipoLogAux['excepcion']


# CRUD
# CREAR
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoArchivos_create(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            request.data['created_at'] = str(timezone_now)
            if 'updated_at' in request.data:
                request.data.pop('updated_at')

            serializer = CreditoArchivosSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def creditoArchivos_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}

            if "minimoCarga" in request.data:
                if request.data["minimoCarga"] != '':
                    filters['fechaCargaArchivo__gte'] = str(request.data["minimoCarga"])

            if "maximoCarga" in request.data:
                if request.data["maximoCarga"] != '':
                    filters['fechaCargaArchivo__lte'] = str(request.data["maximoCarga"])

            if "minimoCreacion" in request.data:
                if request.data["minimoCreacion"] != '':
                    filters['created_at__gte'] = str(request.data["minimoCreacion"])

            if "maximaCreacion" in request.data:
                if request.data["maximaCreacion"] != '':
                    filters['created_at__lte'] = datetime.strptime(request.data['maximaCreacion'],
                                                                   "%Y-%m-%d").date() + timedelta(days=1)

            if "user_id" in request.data:
                if request.data["user_id"] != '':
                    filters['user_id'] = str(request.data["user_id"])

            if "campania" in request.data:
                if request.data["campania"] != '':
                    filters['campania'] = str(request.data["campania"])

            if "empresa_comercial" in request.data:
                if request.data["empresa_comercial"] != '':
                    filters['empresa_comercial'] = str(request.data["empresa_comercial"])

            if "tipoCredito" in request.data:
                if request.data["tipoCredito"] != '':
                    filters['tipoCredito'] = str(request.data["tipoCredito"])

            # Serializar los datos
            query = PreAprobados.objects.filter(**filters).order_by('-created_at')
            serializer = CreditoArchivosSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(), 'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ELIMINAR
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def creditoArchivos_delete(request, pk):
    nowDate = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'delete/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'BORRAR',
        'fechaInicio': str(nowDate),
        'dataEnviada': '{}',
        'fechaFin': str(nowDate),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = PreAprobados.objects.filter(pk=pk, state=1).first()
        except PreAprobados.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
            return Response(status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'DELETE':
            serializer = CreditoArchivosSerializer(query, data={'state': '0', 'updated_at': str(nowDate)}, partial=True)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_200_OK)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)

    # METODO SUBIR ARCHIVOS EXCEL


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_creditosPreaprobados(request, pk):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            archivo = PreAprobados.objects.filter(pk=pk, state=1).first()
            # environ init
            env = environ.Env()
            environ.Env.read_env()  # LEE ARCHIVO .ENV
            client_s3 = boto3.client(
                's3',
                aws_access_key_id=env.str('AWS_ACCESS_KEY_ID'),
                aws_secret_access_key=env.str('AWS_SECRET_ACCESS_KEY')
            )
            with tempfile.TemporaryDirectory() as d:
                ruta = d + 'creditosPreAprobados.xlsx'
                s3 = boto3.resource('s3')
                s3.meta.client.download_file('globalredpymes', str(archivo.linkArchivo), ruta)

            first = True  # si tiene encabezado
            #             uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(ruta)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Clientes"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 9:
                    resultadoInsertar = insertarDato_creditoPreaprobado(dato, archivo.empresa_financiera)
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        os.remove(ruta)
        # archivo.state = 0
        archivo.estado = "Cargado"
        archivo.save()
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# METODO SUBIR ARCHIVOS EXCEL
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_creditosPreaprobados_empleados(request, pk):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            archivo = PreAprobados.objects.filter(pk=pk, state=1).first()
            # environ init
            env = environ.Env()
            environ.Env.read_env()  # LEE ARCHIVO .ENV
            client_s3 = boto3.client(
                's3',
                aws_access_key_id=env.str('AWS_ACCESS_KEY_ID'),
                aws_secret_access_key=env.str('AWS_SECRET_ACCESS_KEY')
            )
            with tempfile.TemporaryDirectory() as d:
                ruta = d + 'creditosPreAprobados.xlsx'
                s3 = boto3.resource('s3')
                s3.meta.client.download_file('globalredpymes', str(archivo.linkArchivo), ruta)

            first = True  # si tiene encabezado
            #             uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(ruta)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Clientes"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 22:
                    resultadoInsertar = insertarDato_creditoPreaprobado_empleado(dato, archivo.empresa_financiera,
                                                                                 archivo.empresa_comercial)
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        os.remove(ruta)
        # archivo.state = 0
        archivo.estado = "Cargado"
        archivo.save()
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_creditoPreaprobado(dato, empresa_financiera):
    try:
        timezone_now = timezone.localtime(timezone.now())
        data = {}
        data['vigencia'] = dato[0].replace('"', "")[0:10] if dato[0] != "NULL" else None
        data['concepto'] = dato[1].replace('"', "") if dato[1] != "NULL" else None
        data['monto'] = dato[2].replace('"', "") if dato[2] != "NULL" else None
        data['plazo'] = dato[3].replace('"', "") if dato[3] != "NULL" else None
        data['interes'] = dato[4].replace('"', "") if dato[4] != "NULL" else None
        data['estado'] = 'PreAprobado'
        data['tipoCredito'] = 'PreAprobado'
        data['canal'] = 'PreAprobado'
        # persona = Personas.objects.filter(identificacion=dato[5],state=1).first()
        # data['user_id'] = persona.user_id
        data['numeroIdentificacion'] = dato[5]
        data['nombres'] = dato[6].replace('"', "") if dato[6] != "NULL" else None
        data['apellidos'] = dato[7].replace('"', "") if dato[7] != "NULL" else None
        data['nombresCompleto'] = data['nombres'] + ' ' + data['apellidos']
        data['empresaIfis_id'] = empresa_financiera
        data['empresasAplican'] = dato[19]
        # Genera el codigo
        codigo = (''.join(random.choice(string.digits) for _ in range(int(6))))
        data['codigoPreaprobado'] = codigo
        data['created_at'] = str(timezone_now)
        # inserto el dato con los campos requeridos
        CreditoPersonas.objects.create(**data)
        subject, from_email, to = 'Crédito de consumo Pre-Aprobado', "08d77fe1da-d09822@inbox.mailtrap.io", \
                                  dato[13]
        txt_content = codigo
        html_content = f"""
        <html>
            <body>
                <h1>PRE-CALIFICACIÓN DE CRÉDITO DE CONSUMO</h1>

                <p>
                 Usted Tiene un crédito Pre-Aprobado de $ {data['monto']} para que realice compras en las mejores Casas Comerciales del país.
                </p>
                <br>
                <p>Ingrese al siguiente link y acceda a su crédito: 
                <a href='{config.API_FRONT_END_CENTRAL}/pages/preApprovedCreditConsumer'>Link</a>
                </p>

                Su código de ingreso es: {codigo}<br>

                Saludos,<br>
                Equipo Global Red Pymes.<br>
            </body>
        </html>
        """
        # CodigoCreditoPreaprobado.objects.create(codigo=codigo, cedula=data['numeroIdentificacion'], monto=data['monto'])
        sendEmail(subject, txt_content, from_email, to, html_content)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_creditoPreaprobado_empleado(dato, empresa_financiera, empresa_comercial):
    print('entro')
    try:
        if (not utils.__validar_ced_ruc(str(dato[7]), 0)):
            return f"""El usuario {dato[8]} {dato[9]} tiene la cédula incorrecta."""

        empleado = Empleados.objects.filter(identificacion=str(dato[7])).first()
        if empleado is None:
            return f"""El usuario {dato[8]} {dato[9]} no está registrado, por favor registre al usuario en la opción EMPLEADOS DE EMPRESAS."""
        else:
            if empleado.empresa._id != ObjectId(empresa_comercial):
                return f"""Lo sentimos, el empleado {dato[8]} {dato[9]} se repite, por favor verifique los datos y vuelva a intentarlo"""

        timezone_now = timezone.localtime(timezone.now())
        data = {}
        data['vigencia'] = dato[0].replace('"', "")[0:10] if dato[0] != "NULL" else None
        data['concepto'] = dato[1].replace('"', "") if dato[1] != "NULL" else None
        data['monto'] = dato[2].replace('"', "") if dato[2] != "NULL" else None
        data['plazo'] = dato[3].replace('"', "") if dato[3] != "NULL" else None
        data['interes'] = dato[4].replace('"', "") if dato[4] != "NULL" else None
        data['estado'] = 'Nuevo'
        data['tipoCredito'] = 'Empleado-PreAprobado'
        data['canal'] = 'Empleado-PreAprobado'
        # persona = Personas.objects.filter(identificacion=dato[5],state=1).first()
        # data['user_id'] = persona.user_id
        data['numeroIdentificacion'] = dato[7]
        data['nombres'] = dato[8].replace('"', "") if dato[8] != "NULL" else None
        data['apellidos'] = dato[9].replace('"', "") if dato[9] != "NULL" else None
        data['nombresCompleto'] = data['nombres'] + ' ' + data['apellidos']
        data['empresaIfis_id'] = empresa_financiera
        data['empresasAplican'] = dato[21]
        # Genera el codigo
        codigo = (''.join(random.choice(string.digits) for _ in range(int(6))))
        data['codigoPreaprobado'] = codigo
        data['created_at'] = str(timezone_now)
        # inserto el dato con los campos requeridos
        CreditoPersonas.objects.create(**data)
        subject, from_email, to = 'Crédito de consumo Pre-Aprobado', "08d77fe1da-d09822@inbox.mailtrap.io", \
                                  dato[15]
        txt_content = codigo
        html_content = f"""
                <html>
                    <body>
                        <h1>PRE-CALIFICACIÓN DE CRÉDITO DE CONSUMO</h1>

                        <p>
                         Usted Tiene un crédito Pre-Aprobado de $ {data['monto']} para que realice compras en las mejores Casas Comerciales del país.
                        </p>
                        <br>
                        <p>Ingrese al siguiente link y acceda a su crédito: 
                        <a href='{config.API_FRONT_END_CENTRAL}/pages/preApprovedCreditConsumer'>Link</a>
                        </p>
        
                        Su código de ingreso es: {codigo}<br>

                        Saludos,<br>
                        Equipo Global Red Pymes.<br>
                    </body>
                </html>
                """
        # CodigoCreditoPreaprobado.objects.create(codigo=codigo, cedula=data['numeroIdentificacion'], monto=data['monto'])
        sendEmail(subject, txt_content, from_email, to, html_content)
        # publish(data)
        return "Dato insertado correctamente"
    except Exception as e:
        return str(e)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_creditosPreaprobados_negocios(request, pk):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            archivo = PreAprobados.objects.filter(pk=pk, state=1).first()
            # environ init
            env = environ.Env()
            environ.Env.read_env()  # LEE ARCHIVO .ENV
            client_s3 = boto3.client(
                's3',
                aws_access_key_id=env.str('AWS_ACCESS_KEY_ID'),
                aws_secret_access_key=env.str('AWS_SECRET_ACCESS_KEY')
            )
            with tempfile.TemporaryDirectory() as d:
                ruta = d + 'creditosPreAprobados.xlsx'
                s3 = boto3.resource('s3')
                s3.meta.client.download_file('globalredpymes', str(archivo.linkArchivo), ruta)

            first = True  # si tiene encabezado
            #             uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(ruta)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Preaprobar líneas-de-crédito_CO"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                if cell.value is None:
                    break
                row_data.append(str(cell.value))
            if row_data:
                lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 14:
                    resultadoInsertar = insertarDato_creditoPreaprobado_microCredito(dato, archivo.empresa_financiera,
                                                                                 archivo.empresa_comercial)
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        os.remove(ruta)
        # archivo.state = 0
        archivo.estado = "Cargado"
        archivo.save()
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_creditoPreaprobado_microCredito(dato, empresa_financiera, empresa_comercial):
    try:
        if (not utils.__validar_ced_ruc(str(dato[8]), 0)):
            return f"""El usuario {dato[5]} {dato[6]} tiene la identificación incorrecta."""

        if (not utils.__validar_ced_ruc(str(dato[4]), 0)):
            return f"""El usuario {dato[3]} tiene el ruc incorrecto."""

        timezone_now = timezone.localtime(timezone.now())
        data = {}
        data['vigencia'] = dato[0].replace('"', "")[0:10] if dato[0] != "NULL" else None
        data['concepto'] = dato[1].replace('"', "") if dato[1] != "NULL" else None
        data['monto'] = dato[2].replace('"', "") if dato[2] != "NULL" else None
        # data['plazo'] = dato[3].replace('"', "") if dato[3] != "NULL" else None
        # data['interes'] = dato[4].replace('"', "") if dato[4] != "NULL" else None
        data['estado'] = 'Nuevo'
        data['tipoCredito'] = 'Pymes-PreAprobado'
        data['canal'] = 'Pymes-PreAprobado'
        # persona = Personas.objects.filter(identificacion=dato[5],state=1).first()
        # data['user_id'] = persona.user_id
        data['numeroIdentificacion'] = dato[8]
        data['nombres'] = dato[5].replace('"', "") if dato[5] != "NULL" else None
        data['apellidos'] = dato[6].replace('"', "") if dato[6] != "NULL" else None
        data['email'] = dato[10].replace('"', "") if dato[10] != "NULL" else None
        data['nombresCompleto'] = data['nombres'] + ' ' + data['apellidos']
        data['empresaIfis_id'] = empresa_financiera
        # data['empresasAplican'] = dato[21]
        # Genera el codigo
        codigo = (''.join(random.choice(string.digits) for _ in range(int(6))))
        data['codigoPreaprobado'] = codigo
        data['created_at'] = str(timezone_now)
        data['razonSocial'] = dato[3]
        data['rucEmpresa'] = dato[4]
        empresaInfo['reprsentante'] = data['nombresCompleto']
        empresaInfo['rucEmpresa'] = dato[4]
        empresaInfo['comercial'] = dato[3]
        empresaInfo['correo'] = data['email']
        empresaInfo['esatdo_civil'] = dato[9]
        empresaInfo['celular'] = dato[11]
        data['empresaInfo'] = empresaInfo
        # inserto el dato con los campos requeridos
        creditoPreAprobado = CreditoPersonas.objects.create(**data)
        creditoSerializer = CreditoPersonasSerializer(creditoPreAprobado)
        subject, from_email, to = 'Usted tiene una LÍNEA DE CRÉDITO PRE-APROBADA PARA PAGAR A SUS PROVEEDORES', "08d77fe1da-d09822@inbox.mailtrap.io", \
                                  dato[10]
        txt_content = codigo
        html_content = f"""
                <html>
                    <body>
                        <p>Estimad@ {data['nombresCompleto']}</p>
                        <br>
                        <p>
                         Nos complace comunicarle que usted tiene una LÍNEA DE CRÉDITO PRE-APROBADA por $ {data['monto']}
                         para que pueda realizar pagos a sus PROVEEDORES con un crédito otorgado {dato[13]}
                        </p>
                        <br>
                        <p>Para acceder a su Línea de Crédito para pago a proveedores haga click en el siguiente enlace:
                        <a href='{config.API_FRONT_END_COOPSANJOSE}/pages/preApprovedCreditLine'>Link</a>
                        </p>

                        Su código de ingreso es: {codigo}
                        <br>
                        <b>Crédito Pagos en la mejor opción de crecimiento para su negocio</b>
                        <br>
                        Saludos,<br>
                        Crédito Pagos – Big Puntos<br>
                    </body>
                </html>
                """
        # CodigoCreditoPreaprobado.objects.create(codigo=codigo, cedula=data['numeroIdentificacion'], monto=data['monto'])
        sendEmail(subject, txt_content, from_email, to, html_content)
        publish(creditoSerializer.data)
        return "Dato insertado correctamente"
    except Exception as e:
        return str(e)


@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def viewEXCEL_creditosPreaprobados_negocios(request, pk):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'GET':
            archivo = PreAprobados.objects.filter(pk=pk, state=1).first()
            # environ init
            env = environ.Env()
            environ.Env.read_env()  # LEE ARCHIVO .ENV
            client_s3 = boto3.client(
                's3',
                aws_access_key_id=env.str('AWS_ACCESS_KEY_ID'),
                aws_secret_access_key=env.str('AWS_SECRET_ACCESS_KEY')
            )
            with tempfile.TemporaryDirectory() as d:
                ruta = d + 'creditosPreAprobados.xlsx'
                s3 = boto3.resource('s3')
                s3.meta.client.download_file('globalredpymes', str(archivo.linkArchivo), ruta)

            first = True  # si tiene encabezado
            #             uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(ruta)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Preaprobar líneas-de-crédito_CO"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                if cell.value is None:
                    break
                row_data.append(str(cell.value))
            if row_data:
                lines.append(row_data)
        os.remove(ruta)
        return Response(lines, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)